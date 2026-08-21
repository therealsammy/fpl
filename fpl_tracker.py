#!/usr/bin/env python3
"""
FPL Season Tracker
==================
Pulls from the official (unofficial) Fantasy Premier League API and builds an
Excel workbook tracking your squad, every player in the game, your gameweek
history, transfers, chips and upcoming fixture difficulty.

It also maintains an append-only snapshot store (fpl_history.csv). The FPL API
only ever reports CURRENT state -- ownership, price, form and expected points
cannot be retrieved retroactively. Every run appends one row per player, so the
store accumulates a real time series you can measure change against.

Run it once a week after the gameweek finishes:

    python fpl_tracker.py

No API key needed. All endpoints are public.

Config: set ENTRY_ID below, or export FPL_ENTRY_ID in the environment.
"""

import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

ENTRY_ID = int(os.environ.get("FPL_ENTRY_ID", 0)) or 8592220   # <-- YOUR manager ID
OUTPUT = Path("fpl_tracker.xlsx")
HISTORY = Path("fpl_history.csv")   # the append-only store. Back this up.
FIXTURE_HORIZON = 6                 # gameweeks ahead for fixture difficulty
TREND_LOOKBACK = 4                  # snapshots back for the medium-term delta
HISTORY_SHEET_LIMIT = 60_000        # cap rows written into the workbook sheet

BASE = "https://fantasy.premierleague.com/api"

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
})

POSITIONS = {1: "GKP", 2: "DEF", 3: "MID", 4: "FWD"}
STATUS = {
    "a": "Available", "d": "Doubtful", "i": "Injured",
    "s": "Suspended", "u": "Unavailable", "n": "Not eligible",
}

# Point-in-time fields. These are the ones the API will never give back.
SNAPSHOT_FIELDS = [
    "Player", "Team", "Pos", "Price", "Owned %", "Form", "Total pts", "GW pts",
    "PPG", "Minutes", "Starts", "Goals", "Assists", "xGI", "DEFCON",
    "DEFCON per 90", "Exp pts next", "Status", "Chance next GW",
    "Transfers in (GW)", "Transfers out (GW)", "Price change (GW)",
]

DELTA_COLS = ("Price", "Owned %", "Form", "Total pts", "xGI")


# ---------------------------------------------------------------------------
# FETCH
# ---------------------------------------------------------------------------

def get(path):
    """GET a JSON endpoint, returning None on 404 rather than raising."""
    r = SESSION.get(f"{BASE}/{path}", timeout=30)
    if r.status_code == 404:
        return None
    r.raise_for_status()
    return r.json()


def fetch_all(entry_id):
    print("Fetching bootstrap-static...")
    boot = get("bootstrap-static/")

    print("Fetching fixtures...")
    fixtures = get("fixtures/")

    print(f"Fetching manager {entry_id}...")
    entry = get(f"entry/{entry_id}/")
    history = get(f"entry/{entry_id}/history/")
    transfers = get(f"entry/{entry_id}/transfers/")

    current_gw, gw_finished = None, False
    for ev in boot["events"]:
        if ev["is_current"]:
            current_gw, gw_finished = ev["id"], ev["finished"]
            break
    if current_gw is None:
        nxt = next((e["id"] for e in boot["events"] if e["is_next"]), 1)
        current_gw = max(1, nxt - 1)

    picks = None
    if entry is not None:
        print(f"Fetching picks for GW{current_gw}...")
        picks = get(f"entry/{entry_id}/event/{current_gw}/picks/")

    return {
        "boot": boot, "fixtures": fixtures, "entry": entry, "history": history,
        "transfers": transfers, "picks": picks,
        "current_gw": current_gw, "gw_finished": gw_finished,
    }


# ---------------------------------------------------------------------------
# TRANSFORM
# ---------------------------------------------------------------------------

def build_players(boot):
    """Every player in the game, flattened with the fields that actually matter."""
    teams = {t["id"]: t["short_name"] for t in boot["teams"]}
    rows = []

    for e in boot["elements"]:
        proj = e.get("price_change_projections") or []
        next_proj = proj[0] if proj else {}

        rows.append({
            "ID": e["id"],
            "Player": e["web_name"],
            "Full name": f"{e['first_name']} {e['second_name']}".strip(),
            "Team": teams.get(e["team"], "?"),
            "Pos": POSITIONS.get(e["element_type"], "?"),
            "Price": e["now_cost"] / 10,
            "Price change (season)": e["cost_change_start"] / 10,
            "Price change (GW)": e["cost_change_event"] / 10,
            "Rise/fall %": _num(next_proj.get("projected_percent")),
            "Change likelihood": next_proj.get("likelihood"),
            "Owned %": _num(e["selected_by_percent"]),
            "Transfers in (GW)": e["transfers_in_event"],
            "Transfers out (GW)": e["transfers_out_event"],
            "Status": STATUS.get(e["status"], e["status"]),
            "Chance next GW": e["chance_of_playing_next_round"],
            "News": e["news"],
            "Total pts": e["total_points"],
            "GW pts": e["event_points"],
            "PPG": _num(e["points_per_game"]),
            "Form": _num(e["form"]),
            "Exp pts next": _num(e["ep_next"]),
            "Value (pts/£m)": _num(e["value_season"]),
            "Minutes": e["minutes"],
            "Starts": e["starts"],
            "Goals": e["goals_scored"],
            "Assists": e["assists"],
            "Clean sheets": e["clean_sheets"],
            "Goals conceded": e["goals_conceded"],
            "Saves": e["saves"],
            "Bonus": e["bonus"],
            "BPS": e["bps"],
            "DEFCON": e["defensive_contribution"],
            "DEFCON per 90": e["defensive_contribution_per_90"],
            "Tackles": e["tackles"],
            "CBI": e["clearances_blocks_interceptions"],
            "Recoveries": e["recoveries"],
            "xG": _num(e["expected_goals"]),
            "xA": _num(e["expected_assists"]),
            "xGI": _num(e["expected_goal_involvements"]),
            "xGC": _num(e["expected_goals_conceded"]),
            "xGI per 90": e["expected_goal_involvements_per_90"],
            "ICT": _num(e["ict_index"]),
            "Yellow": e["yellow_cards"],
            "Red": e["red_cards"],
            "Pens order": e["penalties_order"],
            "Corners/IFK order": e["corners_and_indirect_freekicks_order"],
            "Direct FK order": e["direct_freekicks_order"],
        })

    df = pd.DataFrame(rows)
    return df.sort_values("Total pts", ascending=False).reset_index(drop=True)


# ---------- the time series ----------

def update_history(players_df, gw, gw_finished, path):
    """
    Append this run's snapshot to the store and return the whole series.

    Idempotent on snapshot date: re-running on the same day replaces that
    day's rows rather than duplicating them.
    """
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")

    snap = players_df[["ID"] + SNAPSHOT_FIELDS].copy()
    snap.insert(0, "Snapshot", today)
    snap.insert(1, "Snapshot UTC", now.strftime("%Y-%m-%d %H:%M"))
    snap.insert(2, "GW", gw)
    snap.insert(3, "GW finished", gw_finished)

    if path.exists():
        prior = pd.read_csv(path)
        before = prior["Snapshot"].nunique()
        prior = prior[prior["Snapshot"] != today]
        if prior["Snapshot"].nunique() < before:
            print(f"History: replacing an earlier run from {today}")
        combined = pd.concat([prior, snap], ignore_index=True)
    else:
        combined = snap
        print("History: first snapshot, creating store")

    combined = combined.sort_values(["Snapshot", "ID"]).reset_index(drop=True)
    combined.to_csv(path, index=False)
    print(f"History: {len(combined)} rows across "
          f"{combined['Snapshot'].nunique()} snapshot(s) -> {path}")
    return combined


def _deltas(series, ids, lookback):
    """Latest values plus change vs the previous snapshot and vs `lookback` back."""
    snaps = sorted(series["Snapshot"].unique())
    latest = series[series["Snapshot"] == snaps[-1]].set_index("ID")
    if ids is not None:
        latest = latest[latest.index.isin(ids)]

    out = latest[["Player", "Team", "Pos", "Price", "Owned %", "Form",
                  "Total pts", "xGI", "Exp pts next", "Status"]].copy()

    for label, offset in (("prev", 2), (f"{lookback}w", lookback + 1)):
        past = (series[series["Snapshot"] == snaps[-offset]].set_index("ID")
                if len(snaps) >= offset else None)
        for col in DELTA_COLS:
            if past is None:
                out[f"{col} d{label}"] = None
            else:
                out[f"{col} d{label}"] = (
                    out[col] - past[col].reindex(out.index)
                ).round(2)

    return out.reset_index()


def build_trends(series, squad_ids, lookback):
    """Week-over-week movement for the players you actually own."""
    if series["Snapshot"].nunique() < 2:
        return pd.DataFrame([{
            "Note": "Only one snapshot so far. Trends appear from the second run onward."
        }])
    if not squad_ids:
        return pd.DataFrame([{"Note": "No squad picks available to trend yet."}])
    return _deltas(series, squad_ids, lookback).sort_values("Pos").reset_index(drop=True)


def build_movers(series, lookback):
    """Biggest ownership and price swings across the whole game since last run."""
    if series["Snapshot"].nunique() < 2:
        return pd.DataFrame([{
            "Note": "Only one snapshot so far. Movers appear from the second run onward."
        }])

    d = _deltas(series, None, lookback)
    d = d[d["Owned %"] > 0.5]

    frames = []
    for col, label in (("Owned % dprev", "Ownership"), ("Price dprev", "Price")):
        if col not in d.columns or d[col].isna().all():
            continue
        for direction, asc in (("rising", False), ("falling", True)):
            top = d.sort_values(col, ascending=asc).head(15).copy()
            top.insert(0, "Mover", f"{label} {direction}")
            frames.append(top)

    if not frames:
        return pd.DataFrame([{"Note": "No movement recorded between snapshots."}])

    cols = ["Mover", "Player", "Team", "Pos", "Price", "Price dprev",
            "Owned %", "Owned % dprev", "Form", "Form dprev",
            "Total pts", "Status"]
    out = pd.concat(frames, ignore_index=True)
    return out[[c for c in cols if c in out.columns]]


# ---------- everything else ----------

def build_fixtures(boot, fixtures, horizon):
    """Per-team fixture difficulty grid for the next `horizon` gameweeks."""
    teams = {t["id"]: t["short_name"] for t in boot["teams"]}
    next_gw = next((e["id"] for e in boot["events"] if e["is_next"]), 1)
    window = range(next_gw, next_gw + horizon)

    grid = {short: {"Team": short} for short in teams.values()}

    for f in fixtures:
        gw = f.get("event")
        if gw not in window:
            continue
        h, a = teams.get(f["team_h"]), teams.get(f["team_a"])
        if not h or not a:
            continue
        col = f"GW{gw}"
        grid[h][col] = _append(grid[h].get(col), f"{a} (H) {f['team_h_difficulty']}")
        grid[a][col] = _append(grid[a].get(col), f"{h} (A) {f['team_a_difficulty']}")
        grid[h].setdefault("_fdr", []).append(f["team_h_difficulty"])
        grid[a].setdefault("_fdr", []).append(f["team_a_difficulty"])

    rows = []
    for short, row in grid.items():
        fdrs = row.pop("_fdr", [])
        row["Avg FDR"] = round(sum(fdrs) / len(fdrs), 2) if fdrs else None
        row["Games"] = len(fdrs)
        for gw in window:
            row.setdefault(f"GW{gw}", "BLANK")
        rows.append(row)

    cols = ["Team", "Avg FDR", "Games"] + [f"GW{g}" for g in window]
    return pd.DataFrame(rows)[cols].sort_values("Avg FDR").reset_index(drop=True)


def build_squad(picks, players_df, fixtures_df):
    """Your current 15, joined to live player data and their upcoming run."""
    if not picks:
        return pd.DataFrame([{
            "Note": "No picks available yet. The picks endpoint only returns data "
                    "after a gameweek deadline has passed."
        }])

    by_id = players_df.set_index("ID")
    fdr_by_team = fixtures_df.set_index("Team")["Avg FDR"].to_dict()

    rows = []
    for p in picks["picks"]:
        pid = p["element"]
        if pid not in by_id.index:
            continue
        pl = by_id.loc[pid]
        role = "Captain" if p["is_captain"] else ("Vice" if p["is_vice_captain"] else "")
        rows.append({
            "Slot": p["position"],
            "XI/Bench": "XI" if p["position"] <= 11 else "Bench",
            "Player": pl["Player"],
            "Team": pl["Team"],
            "Pos": pl["Pos"],
            "Role": role,
            "Multiplier": p["multiplier"],
            "Bought at": p.get("purchase_price", 0) / 10 or None,
            "Sell value": p.get("selling_price", 0) / 10 or None,
            "Now": pl["Price"],
            "Profit": round((p.get("selling_price", 0) - p.get("purchase_price", 0)) / 10, 1),
            "Status": pl["Status"],
            "News": pl["News"],
            "Total pts": pl["Total pts"],
            "Form": pl["Form"],
            "Owned %": pl["Owned %"],
            "Exp pts next": pl["Exp pts next"],
            "xGI": pl["xGI"],
            "DEFCON per 90": pl["DEFCON per 90"],
            f"Avg FDR next {FIXTURE_HORIZON}": fdr_by_team.get(pl["Team"]),
            "Rise/fall %": pl["Rise/fall %"],
        })

    return pd.DataFrame(rows).sort_values("Slot").reset_index(drop=True)


def build_season_log(history, boot):
    """Your own gameweek-by-gameweek log (retrievable from the API anytime)."""
    if not history or not history.get("current"):
        return pd.DataFrame([{"Note": "No gameweek history yet."}])

    avg_by_gw = {e["id"]: e["average_entry_score"] for e in boot["events"]}
    rows = []
    for h in history["current"]:
        avg = avg_by_gw.get(h["event"])
        net = h["points"] - h["event_transfers_cost"]
        rows.append({
            "GW": h["event"],
            "Points": h["points"],
            "Hits": -h["event_transfers_cost"],
            "Net": net,
            "Average": avg,
            "vs Average": (net - avg) if avg else None,
            "Total": h["total_points"],
            "GW rank": h["rank"],
            "Overall rank": h["overall_rank"],
            "Transfers": h["event_transfers"],
            "Bench pts": h["points_on_bench"],
            "Squad value": h["value"] / 10,
            "Bank": h["bank"] / 10,
            "Total value": (h["value"] + h["bank"]) / 10,
        })
    return pd.DataFrame(rows)


def build_transfers(transfers, players_df):
    """
    Every transfer you've made.

    Caveat: the 'season pts' columns compare each player's SEASON TOTAL, not
    points accrued since the transfer date. Read the verdict as directional.
    """
    if not transfers:
        return pd.DataFrame([{"Note": "No transfers made yet."}])

    names = players_df.set_index("ID")["Player"].to_dict()
    pts = players_df.set_index("ID")["Total pts"].to_dict()

    rows = []
    for t in reversed(transfers):
        rows.append({
            "GW": t["event"],
            "Date": t["time"][:10],
            "Out": names.get(t["element_out"], t["element_out"]),
            "Sold for": t["element_out_cost"] / 10,
            "Out: season pts": pts.get(t["element_out"]),
            "In": names.get(t["element_in"], t["element_in"]),
            "Bought for": t["element_in_cost"] / 10,
            "In: season pts": pts.get(t["element_in"]),
            "Verdict (season totals)": _verdict(pts.get(t["element_in"]),
                                                pts.get(t["element_out"])),
        })
    return pd.DataFrame(rows)


def build_chips(boot, history):
    """Which chips exist, which you've played, which remain."""
    label = {"wildcard": "Wildcard", "freehit": "Free Hit",
             "bboost": "Bench Boost", "3xc": "Triple Captain"}
    used = {}
    if history and history.get("chips"):
        for c in history["chips"]:
            used.setdefault(c["name"], []).append(c["event"])

    rows = []
    for chip in boot["chips"]:
        name = chip["name"]
        played = [gw for gw in used.get(name, [])
                  if chip["start_event"] <= gw <= chip["stop_event"]]
        rows.append({
            "Chip": label.get(name, name),
            "Half": "First half" if chip["start_event"] < 20 else "Second half",
            "Available": f"GW{chip['start_event']}-{chip['stop_event']}",
            "Status": f"Played GW{played[0]}" if played else "Available",
        })
    return pd.DataFrame(rows)


def build_watchlist(players_df):
    """Top 20 per position by form, filtered to players who actually play."""
    frames = []
    for pos in ["GKP", "DEF", "MID", "FWD"]:
        sub = players_df[(players_df["Pos"] == pos)
                         & (players_df["Status"] == "Available")
                         & (players_df["Minutes"] > 0)].copy()
        sub = sub.sort_values("Form", ascending=False).head(20)
        frames.append(sub[["Player", "Team", "Pos", "Price", "Form", "Total pts",
                           "PPG", "Owned %", "xGI", "DEFCON per 90",
                           "Exp pts next", "Value (pts/£m)"]])
    return pd.concat(frames).reset_index(drop=True)


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _append(existing, new):
    return f"{existing} + {new}" if existing else new


def _verdict(pts_in, pts_out):
    if pts_in is None or pts_out is None:
        return ""
    d = pts_in - pts_out
    return f"Good (+{d})" if d > 0 else (f"Bad ({d})" if d < 0 else "Neutral")


# ---------------------------------------------------------------------------
# WRITE
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)


def write_workbook(sheets, path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

    wb = load_workbook(path)
    for name in wb.sheetnames:
        ws = wb[name]
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY_FONT

        if ws.max_row > 1 and ws.max_column > 1:
            ws.auto_filter.ref = ws.dimensions

        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            width = max((len(str(c.value)) for c in ws[letter]
                         if c.value is not None), default=10)
            ws.column_dimensions[letter].width = min(max(width + 2, 9), 34)

    wb.save(path)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    if ENTRY_ID == 123456:
        print("Set ENTRY_ID at the top of this file, or export FPL_ENTRY_ID.")
        sys.exit(1)

    data = fetch_all(ENTRY_ID)
    boot = data["boot"]

    print("Building sheets...")
    players = build_players(boot)

    series = update_history(players, data["current_gw"],
                            data["gw_finished"], HISTORY)

    fixtures = build_fixtures(boot, data["fixtures"], FIXTURE_HORIZON)
    squad = build_squad(data["picks"], players, fixtures)
    squad_ids = [p["element"] for p in data["picks"]["picks"]] if data["picks"] else []

    trends = build_trends(series, squad_ids, TREND_LOOKBACK)
    movers = build_movers(series, TREND_LOOKBACK)
    season_log = build_season_log(data["history"], boot)
    transfers = build_transfers(data["transfers"], players)
    chips = build_chips(boot, data["history"])
    watchlist = build_watchlist(players)

    snaps = sorted(series["Snapshot"].unique())
    meta = pd.DataFrame([
        {"Field": "Generated", "Value": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")},
        {"Field": "Manager", "Value": (data["entry"] or {}).get("name", "?")},
        {"Field": "Entry ID", "Value": ENTRY_ID},
        {"Field": "Current GW", "Value": data["current_gw"]},
        {"Field": "GW finished", "Value": data["gw_finished"]},
        {"Field": "Overall rank", "Value": (data["entry"] or {}).get("summary_overall_rank")},
        {"Field": "Total points", "Value": (data["entry"] or {}).get("summary_overall_points")},
        {"Field": "Total players", "Value": boot.get("total_players")},
        {"Field": "Snapshots stored", "Value": len(snaps)},
        {"Field": "First snapshot", "Value": snaps[0]},
        {"Field": "Latest snapshot", "Value": snaps[-1]},
        {"Field": "History rows", "Value": len(series)},
        {"Field": "History file", "Value": str(HISTORY)},
        {"Field": "Source", "Value": "fantasy.premierleague.com/api (public, unauthenticated)"},
    ])

    history_sheet = series.tail(HISTORY_SHEET_LIMIT)
    if len(series) > HISTORY_SHEET_LIMIT:
        print(f"Note: History sheet truncated to the most recent "
              f"{HISTORY_SHEET_LIMIT} rows. Full series remains in {HISTORY}.")

    write_workbook({
        "Summary": meta,
        "Squad": squad,
        "Trends": trends,
        "Movers": movers,
        "Season Log": season_log,
        "Chips": chips,
        "Fixtures": fixtures,
        "Watchlist": watchlist,
        "Transfers": transfers,
        "All Players": players,
        "History": history_sheet,
    }, OUTPUT)

    print(f"Written: {OUTPUT.resolve()}")
    print(f"  {len(players)} players | {len(snaps)} snapshot(s) | "
          f"{len(series)} history rows")


if __name__ == "__main__":
    main()


# ---------------------------------------------------------------------------
# NOTES
# ---------------------------------------------------------------------------
#
# THE TWO OUTPUTS
#   fpl_history.csv   Append-only. THIS IS THE IRREPLACEABLE ONE. Ownership,
#                     price and form are point-in-time and the API will not
#                     serve them retroactively. Commit it; back it up.
#   fpl_tracker.xlsx  Disposable. Regenerated from the API plus the store on
#                     every run. Safe to delete.
#
# WHAT NEEDS SNAPSHOTTING AND WHAT DOESN'T
#   Your own results (points, ranks, squad value, transfers, chips) are fully
#   retrievable from entry/{id}/history/ at any time -- no snapshot needed.
#   The player universe is not. That is what the store exists for.
#
# FINDING YOUR ENTRY ID
#   Log in at fantasy.premierleague.com, click "Points". The URL reads
#   .../entry/1234567/event/1 -- the number after /entry/ is your ID.
#
# ENDPOINTS USED (all public, no auth)
#   bootstrap-static/              players, teams, gameweeks, chips, rules
#   fixtures/                      every fixture with FDR ratings
#   entry/{id}/                    manager summary
#   entry/{id}/history/            per-GW history, chips played, past seasons
#   entry/{id}/transfers/          every transfer with prices
#   entry/{id}/event/{gw}/picks/   squad for a gameweek (post-deadline only)
#
# SCHEDULING
#   Tuesday morning: bonus confirmed, prices settled.
#   Cron:    0 8 * * 2  cd /path && /usr/bin/python3 fpl_tracker.py
#   Actions: schedule '0 8 * * 2', then commit fpl_history.csv back to the repo.
#            Put your ID in repo secrets and read it from FPL_ENTRY_ID.
#
# RE-RUNNING
#   Safe. Snapshots are keyed on UTC date, so a second run on the same day
#   replaces that day's rows instead of duplicating them.
#
# DEPENDENCIES
#   pip install -r requirements.txt